import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import cv2
import matplotlib.colors as mpc 
import argparse
import os

from sympy import arg
 
path = os.path.dirname(__file__)
parser = argparse.ArgumentParser()
parser.add_argument('--image', '-i', dest='image_name', default=' ', type=str, nargs='+', help='original image')
parser.add_argument('--output_name', '-o', dest='xlsx_name', default='out.xlsx', type=str, nargs='+', help='the name of xlsx')
parser.add_argument('--shape', '-s', dest='shape', default=' ', type=str, nargs='+', help='the output shape of xlsx')
args=parser.parse_args()


img = cv2.imread(os.path.join(path, args.image_name[0]), cv2.COLOR_BGR2RGB)
if args.image_name[0] == ' ':
    print("Please input the name of the image")
    raise 
if not args.shape[0] == ' ':
    img = cv2.resize(img, [int(args.shape[0][1:-1].split(',')[0]), int(args.shape[0][1:-1].split(',')[1])])
    width = int(args.shape[0][1:-1].split(',')[0])
    length = int(args.shape[0][1:-1].split(',')[1])
else:
    # print("yeah")
    print(img.shape)
    width = img.shape[0]
    length = img.shape[1]
wb = Workbook()
work = wb.active
tmp = []
for i in range(width):
    tmp.append(' ')
tmpp = []
for i in range(length):
    work.append(tmp)

tot = 0
tot_i = 0
for i in img:
    tot_i += 1
    tot_j = -1
    for j in i:
        tot_j += 1
        tot += 1
        if tot % 100 ==0:
            print(tot/(img.size)*3)
        t = j[2]
        j[2] = j[0]
        j[0] = t
        hex_color= mpc.rgb2hex([(x/255) for x in j]) 
        fill = PatternFill('solid',fgColor=str(hex_color)[1:])
        work[tot_i][tot_j].fill = fill
wb.close()
wb.save(os.path.join(path,args.xlsx_name[0]))